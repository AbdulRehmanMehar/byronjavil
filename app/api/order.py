# -*- coding: utf-8 -*-
# app/api/order.py
import mimetypes
from datetime import datetime
import csv, StringIO, xlsxwriter
from openpyxl import load_workbook
from flask import request, make_response, Response
from flask_restplus import Resource, fields
from playhouse.shortcuts import model_to_dict
from sendgrid.helpers.mail import To

from app.server import server
from app.models import OrderState, Order, OrderType, ResearchType, Company, User, ClientCode, UserRole
from app.email import send_mail_template

from .utils import token_required, role_required, order_states
from .geo import geo_address, geo_coordinates

api = server.get_api()
app = server.get_app()
ns = server.get_namespace("supervisor")

order_model = api.model("users_model", {
    'address': fields.String(required=True, description='Property address'),
    'username': fields.String(required=True, description='Assigned Username'),
    'company_id': fields.Integer(required=True, description='Assigned company'),
    'client_code': fields.String(required=True, description='Company client code'),
    'type': fields.String(required=True, description='Order type'),
    'state': fields.String(required=True, description='Order state'),
    'assigned_date': fields.String(required=True, description='Due date'),
    'due_date': fields.String(required=True, description='Due date')
})


@ns.route('/orders')
class SupervisorOrderCollection(Resource):

    @api.doc(security='apikey')
    @token_required
    @role_required(["SUPERVISOR", "SUPERVISOR/MANAGER"])
    def get(self):

        result = list()

        dbo = app.order_dbo
        orders = dbo.read_all()

        for order in orders:
            response = model_to_dict(order)

            response["assigned_date"] = str(response["assigned_date"])
            response["due_date"] = str(response["due_date"])
            if response["state"]:
                response["state"]["state"] = order_states[response["state"]["state"]]
            else:
                response["state"] = {}
                response["state"]["state"] = 'Not Assigned for Research!'

            result.append(response)

        return result

    @ns.expect(order_model)
    @api.doc(security='apikey')
    @token_required
    @role_required(["SUPERVISOR", "SUPERVISOR/MANAGER"])
    def post(self):

        payload = api.payload

        dbo = app.order_dbo
        user_dbo = app.user_dbo
        type_dbo = app.order_type_dbo
        company_dbo = app.company_dbo
        client_dbo = app.client_code_dbo
        use_research = payload["use_research"]
        research_id = payload["research_id"]
        data_id = payload["data_id"]
        company_id = payload["company_id"]
        research_type_id = payload["research"]

        data_user = user_dbo.read_by_id(data_id)
        company = company_dbo.read(company_id)
        order_type = type_dbo.read_by_type(payload["type"])
        client_code = client_dbo.read_by_code(payload["client_code"])

        if research_type_id:
            research_type = ResearchType.select().where(ResearchType.id == research_type_id).get()
        else:
            research_type = None

        if research_id:
            research_user = user_dbo.read_by_id(research_id)
            state = "RESEARCH"

        else:

            state = "DATA_ENTRY"
            research_user = None

        if not use_research:
            research_user = None
            research_type = None

        order_state = OrderState.select().where(OrderState.state == state).get()

        due_date = datetime.strptime(payload["due_date"], "%m/%d/%Y").date()
        assigned_date = datetime.strptime(payload["assigned_date"], "%m/%d/%Y").date()

        full_address = geo_address(payload["address"])
        _lat, _long = geo_coordinates(payload["address"])

        data = {
            "address": payload["address"],
            "full_address": full_address,
            "latitude": _lat,
            "longitude": _long,
            "data_user": data_user,
            "research_type": research_type,
            "research_user": research_user,
            "company": company,
            "due_date": due_date,
            "assigned_date": assigned_date,
            "client_code": client_code,
            "kind": order_type,
            "state": order_state
        }

        order = dbo.create(**data)

        order.save()

        response = model_to_dict(order)

        response["assigned_date"] = str(response["assigned_date"])
        response["due_date"] = str(response["due_date"])
        response["state"]["state"] = order_states[response["state"]["state"]]

        if research_id:
            send_mail_template("ASSIGN_RESEARCH", order, username=research_user.username, from_email="admin@pams.com",
                               to_emails=[research_user.email])
        else:
            send_mail_template("ASSIGN_DATA", order, username=data_user.username, from_email="admin@pams.com",
                               to_emails=[data_user.email])

        return response


@ns.route('/orders/<int:_id>')
class SupervisorOrder(Resource):

    @api.doc(security='apikey')
    @token_required
    @role_required(["SUPERVISOR", "SUPERVISOR/MANAGER"])
    def get(self, _id):

        dbo = app.order_dbo

        order = dbo.read(_id)

        response = model_to_dict(order)

        response["assigned_date"] = str(response["assigned_date"])
        response["due_date"] = str(response["due_date"])
        if response['state']:
            response["state"]["state"] = order_states[response["state"]["state"]]
        else:
            response["state"] = {}
            response["state"]["state"] = 'Not Assigned for Research!'

        return response

    @ns.expect(order_model)
    @api.doc(security='apikey')
    @token_required
    @role_required(["SUPERVISOR", "SUPERVISOR/MANAGER"])
    def put(self, _id):

        payload = api.payload

        dbo = app.order_dbo
        user_dbo = app.user_dbo
        type_dbo = app.order_type_dbo
        company_dbo = app.company_dbo
        client_dbo = app.client_code_dbo
        use_research = payload["use_research"]
        research_type_id = payload["research"]
        research_id = payload["research_id"]
        data_id = payload["data_id"]
        company_id = payload["company_id"]

        if research_type_id:
            research_type = ResearchType.select().where(ResearchType.id == research_type_id).get()
            # research_type =
        else:
            research_type = None

        if research_id:
            research_user = user_dbo.read_by_id(research_id)
            state = "RESEARCH"
            order_state = OrderState.select().where(OrderState.state == state).get()
        else:
            order_state = None
            research_user = None

        if not use_research:
            order_state = None
            research_user = None
            research_type = None

        data_user = user_dbo.read_by_id(data_id)
        company = company_dbo.read(company_id)
        order_type = type_dbo.read_by_type(payload["type"])
        client_code = client_dbo.read_by_code(payload["client_code"])

        full_address = geo_address(payload["address"])
        _lat, _long = geo_coordinates(payload["address"])
        print(research_type)
        data = {
            "address": payload["address"],
            "full_address": full_address,
            "latitude": _lat,
            "longitude": _long,
            "research_type": research_type,
            "research_user": research_user,
            "data_user": data_user,
            "company": company,
            "due_date": payload["due_date"],
            "assigned_date": payload["assigned_date"],
            "client_code": client_code,
            "kind": order_type,
            "state": order_state
        }

        dbo.update(_id, **data)

        order = dbo.read(_id)
        if research_user:
            send_mail_template("ASSIGN_RESEARCH", order, username=research_user.username, from_email="admin@pams.com",
                               to_emails=[research_user.email])

        return True

    @api.doc(security='apikey')
    @token_required
    @role_required(["SUPERVISOR", "SUPERVISOR/MANAGER"])
    def delete(self, _id):

        dbo = app.order_dbo

        dbo.delete(_id)

        return True


@ns.route('/orders/get-csv')
class SupervisorOrderCSV(Resource):
    def get(self):
        si = StringIO.StringIO()
        dbo = app.order_dbo
        orders = dbo.read_all()
        head = [
            "address",
            "full_address",
            "latitude",
            "logintude",
            "assigned_date",
            "due_date",
            "research_completed",
            "data_completed",
            "supervisor_picture",
            "data_picture",
            "manager_submit",
            "company_id",
            "research_user_id",
            "data_user_id",
            "client_code_id",
            "kind_id",
            "research_type_id",
            "state_id"
        ]
        cw = csv.writer(si, dialect="excel")
        cw.writerow(head)

        for order in orders:
            address = order.address
            full_address = order.full_address,
            latitude = order.latitude
            lonitude = order.longitude
            assgined_date = order.assigned_date
            due_date = order.due_date
            research_completed = order.research_completed
            data_completed = order.data_completed
            supervisor_picture = order.supervisor_picture
            data_picture = order.data_picture
            manager_submit = order.manager_submit
            company_id = order.company
            research_user_id = order.research_user
            data_user_id = order.data_user
            client_code_id = order.client_code
            kind_id = order.kind
            research_type_id = order.research_type
            state_id = order.state
            cw.writerow(
                [address, full_address, latitude, lonitude, assgined_date, due_date, research_completed, data_completed,
                 supervisor_picture, data_picture, manager_submit, company_id, research_user_id, data_user_id,
                 client_code_id, kind_id, research_type_id, state_id])

        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = "attachment; filename=export.csv"
        output.headers["Content-type"] = "text/csv"

        return output


def convert_tuple(value):
    return str(value)


@ns.route('/orders/get-xlsx')
class SupervisorOrderXLSX(Resource):
    def get(self):
        dbo = app.order_dbo
        orders = dbo.read_all()
        response = Response()
        response.status_code = 200
        output = StringIO.StringIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('orders')
        bold = workbook.add_format({'bold': 1})

        head = [
            "address",
            "full_address",
            "latitude",
            "logintude",
            "assigned_date",
            "due_date",
            "research_completed",
            "data_completed",
            "supervisor_picture",
            "data_picture",
            "manager_submit",
            "company_id",
            "research_user_id",
            "data_user_id",
            "client_code_id",
            "kind_id",
            "research_type_id",
            "state_id"
        ]

        worksheet.write_row('A1', head, bold)
        idx = 0
        for order in orders:
            idx += 1
            address = order.address
            full_address = order.full_address,
            latitude = order.latitude
            lonitude = order.longitude
            assgined_date = order.assigned_date
            due_date = order.due_date
            research_completed = order.research_completed
            data_completed = order.data_completed
            supervisor_picture = order.supervisor_picture
            data_picture = order.data_picture
            manager_submit = order.manager_submit
            company_id = order.company
            research_user_id = order.research_user
            data_user_id = order.data_user
            client_code_id = order.client_code
            kind_id = order.kind
            research_type_id = order.research_type
            state_id = order.state

            data = [address, full_address, latitude, lonitude, assgined_date, due_date, research_completed,
                    data_completed, supervisor_picture, data_picture, manager_submit, company_id, research_user_id,
                    data_user_id, client_code_id, kind_id, research_type_id, state_id]
            data = map(convert_tuple, data)
            worksheet.write_row(idx, 0, data)

        workbook.close()
        output = make_response(output.getvalue())
        output.headers["Content-Disposition"] = "attachment; filename=export.xlsx"
        output.headers["Content-type"] = "text/xlsx"

        return output


@ns.route('/orders/create/from-xlsx')
class SupervisorCreateOrdersFromXLSX(Resource):
    @api.doc(security='apikey')
    @token_required
    @role_required(["SUPERVISOR", "SUPERVISOR/MANAGER"])
    def post(self):
        dbo = app.order_dbo

        file = request.files['xlsx']
        workbook = load_workbook(file, read_only=True)

        for sheet in workbook.sheetnames:
            row_idx = -1
            for row in workbook[sheet]:
                row_idx += 1
                if row_idx == 0:
                    continue
                else:
                    order_data = []

                    col_idx = 0
                    for col in row:
                        if col_idx <= 8:
                            order_data.append(col.value or None)
                        else:
                            break
                        col_idx += 1

                    print(len(order_data))
                    if len(order_data) == 8: # The last column gets ignored, if there's no data
                        order_data.append(None)

                    if order_data:
                        try:
                            existing_order = Order.select().where(Order.address == order_data[0]).get()
                        except:
                            existing_order = None
                        print (existing_order)

                        if not existing_order:
                            if order_data[8]:
                                research_type = ResearchType.select().where(ResearchType.research_type == order_data[8]).get()
                            else:
                                research_type = None

                            if order_data[4]:
                                research_user = User.select().where(User.username == order_data[4]).get()
                                state = "RESEARCH"
                            else:

                                state = "DATA_ENTRY"
                                research_user = None

                            order_state = OrderState.select().where(OrderState.state == state).get()

                            # due_date = datetime.strptime(str(order_data[2]), "%m-%d-%Y").date()
                            # assigned_date = datetime.strptime(str(order_data[1]), "%m-%d-%Y").date()
                            due_date = str(order_data[2])
                            assigned_date = str(order_data[1])
                            full_address = geo_address(order_data[0])
                            _lat, _long = geo_coordinates(order_data[0])

                            if order_data[3]:
                                company = Company.select().where(Company.name == order_data[3]).get()
                            else:
                                company = None

                            if order_data[5]:
                                data_user = User.select().where(User.username == order_data[5]).get()
                                print (data_user)
                            else:
                                data_user = None

                            if order_data[6]:
                                client_code = ClientCode.select().where(ClientCode.code == order_data[6]).get()
                            else:
                                client_code = None

                            if order_data[7]:
                                kind = OrderType.select().where(OrderType.order_type == order_data[7]).get()
                            else:
                                kind = None

                            data = {
                                "address": order_data[0],
                                "full_address": full_address,
                                "latitude": _lat,
                                "longitude": _long,
                                "data_user": data_user.id if data_user else None,
                                "research_type": research_type.id if research_type else None,
                                "research_user": research_user.id if research_user else None,
                                "company": company.id if company else None,
                                "due_date": due_date,
                                "assigned_date": assigned_date,
                                "client_code": client_code.id if client_code else None,
                                "kind": kind.id if kind else None,
                                "state": order_state
                            }

                            order = dbo.create(**data)

                            order.save()

                            if research_user:
                                send_mail_template("ASSIGN_RESEARCH", order, username=research_user.username,
                                                   from_email="admin@pams.com", to_emails=[research_user.email])
                            else:
                                send_mail_template("ASSIGN_DATA", order, username=data_user.username,
                                                   from_email="admin@pams.com", to_emails=[data_user.email])

                    # else:
                    #     return BaseException("Invalid File")

        return True


@ns.route('/orders/generate/sample-xlsx')
class SupervisorGenerateSampleXLSX(Resource):

    def get(self):
        response = Response()
        response.status_code = 200
        output = StringIO.StringIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('orders')
        bold = workbook.add_format({'bold': 1})
        head = [
            "address",
            "assigned_date",
            "due_date",
            "company",
            "research_user",
            "data_user",
            "client_code",
            "kind",
            "research_type",
        ]

        worksheet.write_row('A1', head, bold)
        try:
            order = Order.select().get()
            try:
                company = Company.select().where(Company.id == order.company).get()
            except:
                company = None
            try:
                research_user = User.select().where(User.id == order.research_user).get()
            except:
                research_user = None

            try:
                data_user = User.select().where(User.id == order.data_user).get()
            except:
                research_user = None

            try:
                client_code = ClientCode.select().where(ClientCode.id == order.client_code).get()
            except:
                client_code = None

            try:
                kind = OrderType.select().where(OrderType.id == order.kind).get()
            except:
                kind = None
            try:
                reseach_type = ResearchType.select().where(ResearchType.id == order.research_type).get()
            except:
                reseach_type = None

            data = [
                order.address,
                str(order.assigned_date),
                str(order.due_date),
                company.name if company else "",
                research_user.username if research_user else "",
                data_user.username if data_user else "",
                client_code.code if client_code else "",
                kind.order_type if kind else "",
                reseach_type.research_type if reseach_type else "",
            ]
            worksheet.write_row(1, 0, data)
        except:
            None

        workbook.close()
        output = make_response(output.getvalue())
        output.headers["Content-Disposition"] = "attachment; filename=sample-upload.xlsx"
        output.headers["Content-type"] = "text/xlsx"

        return output
